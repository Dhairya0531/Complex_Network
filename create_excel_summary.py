import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

output_file = "city_osm_extraction_areas.xlsx"

# 1. Main City OSM Extraction Areas Data
city_extraction_data = [
    {
        "City": "Bengaluru",
        "Country": "India",
        "Topology": "Dense organic, radial arterials",
        "OSM Administrative Boundary": "BBMP (Bruhat Bengaluru Mahanagara Palike)",
        "BBox Width (km)": 35.13,
        "BBox Height (km)": 34.21,
        "Approx. BBox (W x H km)": "35.13 x 34.21",
        "Surface Area (sq km)": 717.21,
        "Lat Range": "[12.8335, 13.1426]",
        "Lon Range": "[77.4599, 77.7841]",
        "Nodes (|V|)": 12846,
        "Edges (|E|)": 28412,
        "Extraction Method": "ox.graph_from_place",
        "Network Type": "drive",
        "Benchmark Sector": "1.5 x 1.5 km (Core Chokepoint)"
    },
    {
        "City": "Berlin",
        "Country": "Germany",
        "Topology": "Grid with central hub corridors",
        "OSM Administrative Boundary": "State/City of Berlin Municipal Boundary",
        "BBox Width (km)": 45.79,
        "BBox Height (km)": 37.74,
        "Approx. BBox (W x H km)": "45.79 x 37.74",
        "Surface Area (sq km)": 890.69,
        "Lat Range": "[52.3382, 52.6755]",
        "Lon Range": "[13.0883, 13.7612]",
        "Nodes (|V|)": 4562,
        "Edges (|E|)": 10234,
        "Extraction Method": "ox.graph_from_place",
        "Network Type": "drive",
        "Benchmark Sector": "1.8 x 1.8 km (Hub Corridors)"
    },
    {
        "City": "London",
        "Country": "UK",
        "Topology": "River-constrained (Thames bridges)",
        "OSM Administrative Boundary": "Greater London Administrative Area",
        "BBox Width (km)": 58.27,
        "BBox Height (km)": 45.09,
        "Approx. BBox (W x H km)": "58.27 x 45.09",
        "Surface Area (sq km)": 1595.49,
        "Lat Range": "[51.2868, 51.6919]",
        "Lon Range": "[-0.5104, 0.3340]",
        "Nodes (|V|)": 9204,
        "Edges (|E|)": 21186,
        "Extraction Method": "ox.graph_from_place",
        "Network Type": "drive",
        "Benchmark Sector": "1.6 x 1.6 km (Thames Bridges)"
    },
    {
        "City": "Sydney",
        "Country": "Australia",
        "Topology": "Harbor-constrained, bridge bottlenecks",
        "OSM Administrative Boundary": "Greater Sydney Metropolitan Extent",
        "BBox Width (km)": 100.22,
        "BBox Height (km)": 90.05,
        "Approx. BBox (W x H km)": "100.22 x 90.05",
        "Surface Area (sq km)": 4367.61,
        "Lat Range": "[-34.1732, -33.3642]",
        "Lon Range": "[150.2608, 151.3439]",
        "Nodes (|V|)": 11284,
        "Edges (|E|)": 24618,
        "Extraction Method": "ox.graph_from_place",
        "Network Type": "drive",
        "Benchmark Sector": "2.0 x 2.0 km (Harbor Bottlenecks)"
    },
    {
        "City": "Nancy",
        "Country": "France",
        "Topology": "Radial hub and spoke arterial",
        "OSM Administrative Boundary": "Nancy Municipality Boundary",
        "BBox Width (km)": 5.69,
        "BBox Height (km)": 4.67,
        "Approx. BBox (W x H km)": "5.69 x 4.67",
        "Surface Area (sq km)": 14.97,
        "Lat Range": "[48.6669, 48.7092]",
        "Lon Range": "[6.1342, 6.2126]",
        "Nodes (|V|)": 1240,
        "Edges (|E|)": 2810,
        "Extraction Method": "ox.graph_from_place",
        "Network Type": "drive",
        "Benchmark Sector": "1.2 x 1.2 km (Radial Hub)"
    }
]

# 2. Simulation Parameters Data
sim_params_data = [
    {"Parameter Name": "Simulation Steps (T)", "Symbol": "T", "Value": 90, "Unit": "steps", "Description": "Total discrete simulation steps per trial"},
    {"Parameter Name": "Signal Cycle Time", "Symbol": "T_cyc", "Value": 60, "Unit": "seconds", "Description": "Traffic signal cycle length"},
    {"Parameter Name": "Minimum Green Time", "Symbol": "g_min", "Value": 15, "Unit": "seconds", "Description": "Lower bound on dynamically allocated green time"},
    {"Parameter Name": "Maximum Green Time", "Symbol": "g_max", "Value": 45, "Unit": "seconds", "Description": "Upper bound on dynamically allocated green time"},
    {"Parameter Name": "Baseline Arrival Rate", "Symbol": "lambda", "Value": 200, "Unit": "veh/step", "Description": "Poisson mean vehicle arrival rate"},
    {"Parameter Name": "OD Routes per City", "Symbol": "R", "Value": 20, "Unit": "pairs", "Description": "Origin-Destination route bank through structural bottlenecks"},
    {"Parameter Name": "Betweenness Sample Size", "Symbol": "k", "Value": 80, "Unit": "sources", "Description": "Source nodes sampled for Brandes betweenness approximation"},
    {"Parameter Name": "Evaluation Trials", "Symbol": "N_trials", "Value": 20, "Unit": "trials", "Description": "Randomized Monte Carlo evaluation trials per controller"},
    {"Parameter Name": "Saturation Lane Capacity", "Symbol": "C_lane", "Value": 8, "Unit": "veh/cycle/lane", "Description": "Discharge capacity standard (1800 veh/hr/lane)"},
    {"Parameter Name": "Centrality Scaling Factor", "Symbol": "centrality_scale", "Value": 2.0, "Unit": "multiplier", "Description": "Weight multiplier for structural importance score"}
]

# Write to Excel with styling
wb = openpyxl.Workbook()

# Sheet 1: City OSM Extraction Areas
ws1 = wb.active
ws1.title = "City Extraction Areas"
ws1.views.sheetView[0].showGridLines = True

# Style definitions
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
accent_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
border_thin = Side(border_style="thin", color="D3D3D3")
cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
title_font = Font(name="Calibri", size=14, bold=True, color="1F4E79")

# Title
ws1.append(["OpenStreetMap City Extraction Areas and Network Topology Properties"])
ws1.merge_cells("A1:N1")
ws1["A1"].font = title_font
ws1["A1"].alignment = Alignment(vertical="center")
ws1.row_dimensions[1].height = 30
ws1.append([]) # empty row

df1 = pd.DataFrame(city_extraction_data)
headers1 = list(df1.columns)
ws1.append(headers1)

for r_idx, row in enumerate(df1.itertuples(index=False), start=4):
    ws1.append(list(row))
    ws1.row_dimensions[r_idx].height = 22
    for c_idx in range(1, len(headers1) + 1):
        cell = ws1.cell(row=r_idx, column=c_idx)
        cell.border = cell_border
        cell.font = Font(name="Calibri", size=10.5)
        cell.fill = accent_fill if r_idx % 2 == 0 else white_fill
        if headers1[c_idx - 1] in ["BBox Width (km)", "BBox Height (km)", "Surface Area (sq km)"]:
            cell.number_format = "#,##0.00"
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif headers1[c_idx - 1] in ["Nodes (|V|)", "Edges (|E|)"]:
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif headers1[c_idx - 1] in ["Lat Range", "Lon Range", "Approx. BBox (W x H km)"]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")

# Format header row
ws1.row_dimensions[3].height = 28
for c_idx in range(1, len(headers1) + 1):
    cell = ws1.cell(row=3, column=c_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = cell_border

# Auto-adjust column widths
for col in ws1.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws1.column_dimensions[col_letter].width = max(max_len + 4, 12)

# Sheet 2: Simulation Parameters
ws2 = wb.create_sheet(title="Simulation Parameters")
ws2.views.sheetView[0].showGridLines = True

ws2.append(["UTSON Multi-City Simulation Parameters & Experimental Constants"])
ws2.merge_cells("A1:E1")
ws2["A1"].font = title_font
ws2["A1"].alignment = Alignment(vertical="center")
ws2.row_dimensions[1].height = 30
ws2.append([])

df2 = pd.DataFrame(sim_params_data)
headers2 = list(df2.columns)
ws2.append(headers2)

for r_idx, row in enumerate(df2.itertuples(index=False), start=4):
    ws2.append(list(row))
    ws2.row_dimensions[r_idx].height = 22
    for c_idx in range(1, len(headers2) + 1):
        cell = ws2.cell(row=r_idx, column=c_idx)
        cell.border = cell_border
        cell.font = Font(name="Calibri", size=10.5)
        cell.fill = accent_fill if r_idx % 2 == 0 else white_fill
        if headers2[c_idx - 1] == "Value":
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif headers2[c_idx - 1] in ["Symbol", "Unit"]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")

ws2.row_dimensions[3].height = 28
for c_idx in range(1, len(headers2) + 1):
    cell = ws2.cell(row=3, column=c_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = cell_border

for col in ws2.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws2.column_dimensions[col_letter].width = max(max_len + 4, 12)

wb.save(output_file)
print(f"Successfully generated: {output_file}")
